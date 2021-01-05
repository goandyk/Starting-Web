from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭생상 변경
# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet, 2")

wb.save("sample1.xlsx")
